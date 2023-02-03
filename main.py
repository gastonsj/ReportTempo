import json
import pandas as pd
import numpy as np
from datetime import datetime
from calendar import monthrange
from openpyxl import Workbook
import myFunctions as fx

# JSON file
f = open ('config.ini', "r")
jsondata = json.loads(f.read())

token_jira=jsondata['token_jira']
email_jira=jsondata['email_jira']
token_tempo=jsondata['token_tempo']
query_month=jsondata['query_month']
query_year=jsondata['query_year']

num_days = monthrange(int(query_year), int(query_month))[1]
working_date=datetime.strptime(query_month+'-'+query_year, '%m-%Y')
query_data=True
if query_data:
    team_data=fx.query_team(query_year+'-'+query_month+'-01',query_year+'-'+query_month+'-'+str(num_days),token_tempo)
    issue_id=[]
    timeSpent=[]
    startDate=[]
    startTime=[]
    author_id=[]
    for r in team_data['results']:
        issue_id.append(r['issue']['id'])
        timeSpent.append(r['timeSpentSeconds'])
        startDate.append(r['startDate'])
        startTime.append(r['startTime'])
        author_id.append(r['author']['accountId'])

    issue_distinct=fx.distinctList(issue_id)
    issues=[]
    for i in issue_distinct:
        issues.append(fx.query_issue(i,email_jira,token_jira))
    issue_dic = {issue_distinct[i]: issues[i] for i in range(len(issue_distinct))}
    issue_name=[]
    issue_key=[]
    for i in issue_id:
        issue_key.append(issue_dic[i]['key'])
        issue_name.append(issue_dic[i]['name'])

    author_distinct=fx.distinctList(author_id)
    author_name=[]
    for i in author_distinct:
        author_name.append(fx.query_author(i,email_jira,token_jira))
    author_dic = {author_distinct[i]: author_name[i] for i in range(len(author_distinct))}
    author_name=[]
    for i in author_id:
        author_name.append(author_dic[i])
    df = pd.DataFrame({
        'Issue ID':issue_id,
        'Issue Key':issue_key,
        'Issue Name': issue_name,
        'Time Spent': timeSpent,
        'Start Date': startDate,
        'Start Time': startTime,
        'Author ID': author_id,
        'Author Name': author_name
    })
    # Write DataFrame to CSV File with Default params.
    df.to_csv("df.csv")
else:
    df = pd.read_csv('df.csv')

domingos=[]
for i in range(num_days-1):
    dt1=datetime.strptime(query_year+'-'+query_month+'-'+str(i+1), '%Y-%m-%d').isocalendar()[1]
    dt2=datetime.strptime(query_year+'-'+query_month+'-'+str(i+2), '%Y-%m-%d').isocalendar()[1]
    if dt1!=dt2:
        domingos.append(i+1)
dic_semanas={}
for i in range(num_days):
    dt1=datetime.strptime(query_year+'-'+query_month+'-'+str(i+1), '%Y-%m-%d').isocalendar()[1]
    if dt1 not in dic_semanas:
        dic_semanas.update({dt1:len(dic_semanas)})
semanas=[]
if domingos[0]!=1:
    if domingos[0]==2:
        semanas.append('1') 
    else: 
        semanas.append(str(1)+'-'+str(domingos[0]-1))
for i in range(len(domingos)-1):
    semanas.append(str(domingos[i])+'-'+str(domingos[i+1]-1))
semanas.append(str(domingos[-1])+'-'+str(num_days))


workbook = Workbook()
#SHEET 1: FACTURABLES
sheet1=workbook[workbook.sheetnames[0]]
sheet1.title ='FACTURABLES'
#Encabezado
sheet1["A1"] = "Proyecto"
sheet1["B1"] = "Clave"
sheet1["C1"] = "Usuarios"
for i in range(len(semanas)):
    sheet1.cell(row=1, column=i+4).value=datetime.strftime(working_date,'%b/%Y ') + semanas[i]
#Proyectos + horas
author_distinct=fx.distinctList(df['Author Name'])
issue_name_distinct=fx.distinctList(df['Issue Name'])
issue_key_distict=fx.distinctList(df['Issue Key'])
issues_dic_complete={}
for i in range(len(issue_name_distinct)):
    issues_dic_complete.update({issue_key_distict[i]:issue_name_distinct[i]})
actual_row=2
for i in issues_dic_complete.keys():
    if i != 'CLOUDDT':
        sheet1.cell(actual_row, column=1).value=issues_dic_complete[i]
        sheet1.cell(actual_row, column=2).value=i
        dic_count_hour=fx.dic_count_hour_empty(len(semanas),author_distinct)
        #print(dic_count_hour)
        for j in range(len(df)):
            if df['Issue Key'][j]==i:
                date_isocalendar=datetime.strptime(df['Start Date'][j], '%Y-%m-%d').isocalendar()[1]
                n=df['Author Name'][j]
                d=dic_semanas[date_isocalendar]-1
                dic_count_hour[n][d]+=df['Time Spent'][j]/60/60
        top_row=actual_row
        actual_row+=1
        for k in author_distinct:
            if fx.sum_list(dic_count_hour[k])!=0:
                sheet1.cell(actual_row, column=3).value=k
                for j in range(len(semanas)):
                    sheet1.cell(actual_row, column=j+4).value=dic_count_hour[k][j]
                actual_row+=1
        for j in range(len(semanas)):
            suma=0
            for k in range(top_row+1,actual_row):
                suma+=sheet1.cell(k, column=j+4).value
            sheet1.cell(top_row, column=j+4).value=suma 
sheet1.cell(actual_row, column=4+len(semanas)).value="TOTALES: "
sheet1.cell(actual_row, column=5+len(semanas)).value="[%]"
actual_row+=1     
sheet1.cell(actual_row, column=1).value="FACTURABLE: "
sheet1.cell(actual_row+1, column=1).value="NO FACTURABLE: "
sheet1.cell(actual_row+2, column=1).value="TOTAL: "
suma_facturables=[0 for i in range(len(semanas))]
suma_nofacturables=[0 for i in range(len(semanas))]
for i in range(len(df)):
    date_isocalendar=datetime.strptime(df['Start Date'][i], '%Y-%m-%d').isocalendar()[1]
    if df['Issue Key'][i]=='CLOUDDT':   
        suma_nofacturables[dic_semanas[date_isocalendar]-1]+=df['Time Spent'][i]/60/60
    else:
        suma_facturables[dic_semanas[date_isocalendar]-1]+=df['Time Spent'][i]/60/60
for i in range(len(semanas)):
    sheet1.cell(actual_row, column=4+i).value=suma_facturables[i]
    sheet1.cell(actual_row+1, column=4+i).value=suma_nofacturables[i]
    sheet1.cell(actual_row+2, column=4+i).value=suma_facturables[i]+suma_nofacturables[i]
sheet1.cell(actual_row, column=4+len(semanas)).value=fx.sum_list(suma_facturables)
sheet1.cell(actual_row+1, column=4+len(semanas)).value=fx.sum_list(suma_nofacturables)
sheet1.cell(actual_row+2, column=4+len(semanas)).value=fx.sum_list(suma_facturables)+fx.sum_list(suma_nofacturables)
total=fx.sum_list(suma_facturables)+fx.sum_list(suma_nofacturables)
if total>0:
    sheet1.cell(actual_row, column=5+len(semanas)).value=round(fx.sum_list(suma_facturables)/total*100,2)
    sheet1.cell(actual_row+1, column=5+len(semanas)).value=round(fx.sum_list(suma_nofacturables)/total*100,2)
#SHEET 2: CLOUDDT
sheet2 = workbook.create_sheet("CLOUDDT")
#Encabezado
sheet2["A1"] = "Proyecto"
sheet2["B1"] = "Clave"
sheet2["C1"] = "Usuarios"
for i in range(len(semanas)):
    sheet2.cell(row=1, column=i+4).value=datetime.strftime(working_date,'%b/%Y ') + semanas[i]  
#Proyectos + horas
actual_row=2
sheet2.cell(actual_row, column=1).value=issues_dic_complete['CLOUDDT']
sheet2.cell(actual_row, column=2).value='CLOUDDT'
dic_count_hour=fx.dic_count_hour_empty(len(semanas),author_distinct)
for i in range(len(df)):
    if df['Issue Key'][i]=='CLOUDDT':
        date_isocalendar=datetime.strptime(df['Start Date'][i], '%Y-%m-%d').isocalendar()[1]
        n=df['Author Name'][i]
        d=dic_semanas[date_isocalendar]-1
        dic_count_hour[n][d]+=df['Time Spent'][i]/60/60
top_row=actual_row
actual_row+=1
for k in author_distinct:
    if fx.sum_list(dic_count_hour[k])!=0:
        sheet2.cell(actual_row, column=3).value=k
        for j in range(len(semanas)):
            sheet2.cell(actual_row, column=j+4).value=dic_count_hour[k][j]
        actual_row+=1
for j in range(len(semanas)):
    suma=0
    for k in range(top_row+1,actual_row):
        suma+=sheet2.cell(k, column=j+4).value
    sheet2.cell(top_row, column=j+4).value=suma 

filenamexls='reporte-'+ datetime.strftime(working_date,'%b%y') + '.xlsx'
workbook.save(filenamexls)